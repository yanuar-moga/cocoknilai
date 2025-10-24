# app.py
"""
MB Tools - Pencocokan Nilai Siswa (Streamlit Version)
Converted by ChatGPT for Bayu (SMPN1Moga)
- Logic core dipertahankan (match_and_write)
- GUI: Streamlit
"""

import os
import time
import tempfile
import io
from datetime import datetime

import streamlit as st
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------
# Helpers (same logic as original)
# ---------------------------------------------------------
def parse_score(raw):
    if pd.isna(raw):
        return None
    s = str(raw).strip()
    if "/" in s:
        s = s.split("/")[0].strip()
    if s.endswith("%"):
        s = s[:-1].strip()
    cleaned = "".join(ch for ch in s if (ch.isdigit() or ch in ".,")) 
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except:
        return None

def normalize_text(s):
    if pd.isna(s):
        return ""
    return " ".join(str(s).strip().lower().split())

def cari_kolom_otomatis(df, keywords):
    cols = list(df.columns)
    for kw in keywords:
        for col in cols:
            name = str(col).strip().lower()
            if kw in name:
                return col
    return None

def apply_color(cell, value):
    if value is None:
        return
    try:
        val = float(value)
    except:
        return
    if val < 78:
        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # merah
    elif val < 80:
        cell.fill = PatternFill(start_color="FFF58C", end_color="FFF58C", fill_type="solid")  # kuning
    else:
        cell.fill = PatternFill(start_color="B7E1A1", end_color="B7E1A1", fill_type="solid")  # hijau

# ---------------------------------------------------------
# Core matching logic (accepts file paths)
# ---------------------------------------------------------
def match_and_write(respons_path, hasil_path, log_fn=None, progress_fn=None):
    def log(msg):
        if log_fn:
            log_fn(msg)
        else:
            print(msg)

    log("📖 Membaca file Excel...")
    df_resp = pd.read_excel(respons_path, engine="openpyxl")
    df_hasil = pd.read_excel(hasil_path, engine="openpyxl")

    # deteksi kolom
    name_keys = ["nama", "name"]
    score_keys = ["score", "nilai", "skor"]
    absen_keys = ["absen", "no", "nomor", "nis", "id"]
    time_keys = ["time", "timestamp", "tgl", "waktu"]

    # safe fallback to avoid index errors
    def col_or(index, df):
        try:
            return df.columns[index]
        except:
            return df.columns[0]

    col_name = cari_kolom_otomatis(df_resp, name_keys) or col_or(2, df_resp)
    col_score = cari_kolom_otomatis(df_resp, score_keys) or col_or(1, df_resp)
    col_time = cari_kolom_otomatis(df_resp, time_keys) or col_or(0, df_resp)
    col_absen = cari_kolom_otomatis(df_resp, absen_keys)

    kolom_nama_hasil = cari_kolom_otomatis(df_hasil, name_keys) or col_or(1, df_hasil)
    kolom_absen_hasil = cari_kolom_otomatis(df_hasil, absen_keys) or col_or(0, df_hasil)

    # normalize
    df_resp["_name_norm"] = df_resp[col_name].apply(normalize_text)
    df_hasil["_name_norm"] = df_hasil[kolom_nama_hasil].apply(normalize_text)
    if col_absen in df_resp.columns:
        df_resp["_absen_str"] = df_resp[col_absen].astype(str).fillna("").str.strip()
    else:
        df_resp["_absen_str"] = ""
    df_hasil["_absen_str"] = df_hasil[kolom_absen_hasil].astype(str).fillna("").str.strip()

    # siapkan kolom Score_1..6
    for i in range(1, 7):
        colname = f"Score_{i}"
        if colname not in df_hasil.columns:
            df_hasil[colname] = pd.NA

    total_resp = len(df_resp)
    processed = 0

    # proses pencocokan
    log("🔁 Mencocokkan data siswa...")
    for ridx, rrow in df_resp.iterrows():
        processed += 1
        if progress_fn:
            progress_fn(int(processed / max(total_resp, 1) * 100))

        raw_name = rrow["_name_norm"]
        raw_absen = str(rrow["_absen_str"])
        raw_score = parse_score(rrow[col_score])

        matched_idx = None

        # 1. nama cocok langsung
        for hid, target_norm in df_hasil["_name_norm"].items():
            if raw_name == target_norm or raw_name in target_norm or target_norm in raw_name:
                matched_idx = hid
                break

        # 2. cocok absen
        if matched_idx is None and raw_absen:
            for hid, a in df_hasil["_absen_str"].items():
                if str(a).strip() == raw_absen.strip():
                    matched_idx = hid
                    break

        # 3. fuzzy match
        if matched_idx is None and raw_name:
            best_score, best_idx = 0, None
            for hid, target_norm in df_hasil["_name_norm"].items():
                sc = fuzz.token_set_ratio(raw_name, target_norm)
                if sc > best_score:
                    best_score, best_idx = sc, hid
            if best_score >= 65:
                matched_idx = best_idx

        if matched_idx is None:
            log(f"⚠️ Tidak ditemukan: {rrow[col_name]}")
            continue

        # Isi ke kolom Score_1–6
        for i in range(1, 7):
            coln = f"Score_{i}"
            val = df_hasil.at[matched_idx, coln]
            if pd.isna(val) or str(val).strip() == "":
                df_hasil.at[matched_idx, coln] = raw_score
                break

    # -----------------------------
    # Hitung kolom SCORE otomatis
    # -----------------------------
    log("🧮 Menghitung kolom SCORE...")
    def safe_float(x):
        try:
            return float(x)
        except:
            return None

    def hitung_score(row):
        skor_list = [safe_float(row.get(f"Score_{i}")) for i in range(1, 7)]
        s1 = skor_list[0]
        if s1 and s1 >= 80:
            return s1
        elif any(s and s >= 80 for s in skor_list[1:]):
            return 78
        return None

    df_hasil["SCORE"] = df_hasil.apply(hitung_score, axis=1)

    # simpan hasil
    out_path = os.path.join(os.path.dirname(os.path.abspath(hasil_path)), "hasil_pencocokan.xlsx")
    df_hasil.drop(columns=["_name_norm", "_absen_str"], inplace=True, errors="ignore")
    df_hasil.to_excel(out_path, index=False)

    # Pewarnaan skor otomatis (openpyxl)
    wb = load_workbook(out_path)
    ws = wb.active
    # cari kolom Score_1..Score_6 dalam sheet (asumsi mulai dari kolom mana pun)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # cari index kolom Score_1..6
    score_cols_idx = []
    for i, h in enumerate(headers, start=1):
        if h and isinstance(h, str) and h.strip().startswith("Score_"):
            score_cols_idx.append(i)
    # apply warna ke setiap row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in score_cols_idx:
            cell = row[idx - 1]
            apply_color(cell, cell.value)
    wb.save(out_path)

    log(f"✅ Selesai! File disimpan di: {out_path}")
    if progress_fn:
        progress_fn(100)
    return out_path

# ---------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------
st.set_page_config(page_title="MB Tools — Pencocokan Nilai Siswa", layout="wide")

# Header
st.markdown(
    """
    <div style="background:#1f6feb;padding:18px;border-radius:8px">
      <h2 style="color:white;margin:0">📊 MB Tools — Pencocokan Nilai Siswa (Streamlit)</h2>
      <div style="color:#e6f0ff">Credit: Apps by MB — Donasi: wa.me/628522939579</div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.write("")  # spacer

# Two-column uploader layout
col1, col2 = st.columns([1, 1])

with col1:
    st.subheader("📂 File Respons (upload)")
    uploaded_respons = st.file_uploader("Pilih file respons (Excel)", type=["xlsx", "xls"], key="resp")
    if uploaded_respons:
        try:
            df_preview_resp = pd.read_excel(uploaded_respons, engine="openpyxl", nrows=5)
            st.markdown("**Preview (5 baris pertama)**")
            st.dataframe(df_preview_resp)
        except Exception as e:
            st.warning(f"Gagal membaca preview respons: {e}")

with col2:
    st.subheader("📘 File Hasil (upload)")
    uploaded_hasil = st.file_uploader("Pilih file hasil (Excel) — template kelas", type=["xlsx", "xls"], key="hasil")
    if uploaded_hasil:
        try:
            df_preview_hasil = pd.read_excel(uploaded_hasil, engine="openpyxl", nrows=5)
            st.markdown("**Preview (5 baris pertama)**")
            st.dataframe(df_preview_hasil)
        except Exception as e:
            st.warning(f"Gagal membaca preview hasil: {e}")

st.write("")  # spacer

# Options and action
col_a, col_b, col_c = st.columns([1, 1, 2])
with col_a:
    auto_rename = st.checkbox("Auto-rename output with timestamp", value=True)
with col_b:
    show_log_detail = st.checkbox("Tampilkan log detail", value=True)

with col_c:
    st.write("")  # align
    process_btn = st.button("🚀 Proses Data", use_container_width=True)

# Log & progress UI placeholders
log_box = st.empty()
progress_container = st.empty()

# Helper to save uploaded file to temp path
def save_uploaded_to_temp(uploaded_file, prefix):
    if uploaded_file is None:
        return None
    # ensure temp dir
    tmpdir = tempfile.mkdtemp(prefix="mbtools_")
    ext = os.path.splitext(uploaded_file.name)[1] or ".xlsx"
    tmp_path = os.path.join(tmpdir, prefix + ext)
    # write bytes
    with open(tmp_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return tmp_path

# Run process when button clicked
if process_btn:
    if not uploaded_respons or not uploaded_hasil:
        st.warning("Silakan upload kedua file (Respons & Hasil) terlebih dahulu.")
    else:
        logs = []
        def log_fn(msg):
            # append timestamp
            ts = datetime.now().strftime("%H:%M:%S")
            entry = f"[{ts}] {msg}"
            logs.append(entry)
            if show_log_detail:
                log_box.text("\n".join(logs[-200:]))
            else:
                # only show last 6 entries
                log_box.text("\n".join(logs[-6:]))

        prog_bar = progress_container.progress(0)
        pct_text = progress_container.empty()

        def progress_fn(p):
            try:
                prog_bar.progress(p)
                pct_text.text(f"{p}%")
            except Exception:
                pass

        # Save uploaded files to disk so match_and_write can open them
        try:
            path_resp = save_uploaded_to_temp(uploaded_respons, "respons")
            path_hasil = save_uploaded_to_temp(uploaded_hasil, "hasil")
            if auto_rename:
                # ensure output goes to the hasil temp folder
                # match_and_write will write hasil_pencocokan.xlsx in same dir as hasil_path
                out_dir = os.path.dirname(path_hasil)
                # call
                out_path = match_and_write(path_resp, path_hasil, log_fn=log_fn, progress_fn=progress_fn)
            else:
                out_path = match_and_write(path_resp, path_hasil, log_fn=log_fn, progress_fn=progress_fn)
        except Exception as e:
            st.error(f"❌ Terjadi error saat memproses: {e}")
            log_fn(f"ERROR: {e}")
            prog_bar.progress(0)
            pct_text.text("")
        else:
            st.success("✅ Proses selesai!")
            log_fn(f"File output: {out_path}")
            # read file bytes for download
            try:
                with open(out_path, "rb") as f:
                    data = f.read()
                # create a nicer filename for download
                fname = "hasil_pencocokan.xlsx"
                if auto_rename:
                    tstamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    fname = f"hasil_pencocokan_{tstamp}.xlsx"
                st.download_button("⬇️ Download Hasil Pencocokan", data, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.warning(f"Gagal menyediakan download: {e}")

        # finally show last logs
        if show_log_detail:
            log_box.text("\n".join(logs[-200:]))
        else:
            log_box.text("\n".join(logs[-20:]))

st.write("")  # bottom spacer
st.markdown("<small>Apps by MB — Donasi/Support: wa.me/628522939579</small>", unsafe_allow_html=True)
